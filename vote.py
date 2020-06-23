import asyncio
import discord
import datetime
import openpyxl
import random
client = discord.Client()
# 1-6에서 생성된 토큰을 이곳에 입력해주세요.
token = ""

# 봇이 구동되었을 때 동작되는 코드입니다.
@client.event
async def on_ready():
    print("Logged in as ") #화면에 봇의 아이디, 닉네임이 출력됩니다.
    print(client.user.name)
    print(client.user.id)
    print("===========")
    # 디스코드에는 현재 본인이 어떤 게임을 플레이하는지 보여주는 기능이 있습니다.
    # 이 기능을 사용하여 봇의 상태를 간단하게 출력해줄 수 있습니다.
    await client.change_presence(game=discord.Game(name="투표", type=1))

# 봇이 새로운 메시지를 수신했을때 동작되는 코드입니다.


@client.event
async def on_message(message):
    if message.author.bot: #만약 메시지를 보낸사람이 봇일 경우에는
        return None #동작하지 않고 무시합니다.

    id = message.author.id #id라는 변수에는 메시지를 보낸사람의 ID를 담습니다.
    channel = message.channel #channel이라는 변수에는 메시지를 받은 채널의 ID를 담습니다.

    #투표설정
    if message.content.startswith('!투표설정'):
        order=message.content[6]
        
        file_parti=openpyxl.load_workbook("참여자.xlsx")
        file_vote=openpyxl.load_workbook("투표목록.xlsx")
        sheet_parti=file_parti.active
        sheet_vote=file_vote.active

        flag=True
        flag_preEnd=False
        if sheet_parti["A1"].value!=0:
            flag_preEnd=True
        if order=='0':
            sheet_parti["A1"]=1
        elif order=='1':
            sheet_parti["A1"]=2
        else:
            await client.send_message(channel,'잘못된 입력 형식입니다.')
            flag=False
            

        if flag:
            #이전에 진행중이었던 투표가 있었다면 결과 출력
            if flag_preEnd:
                await client.send_message(channel,'이전의 투표를 종료합니다.')
                flag=True
                if sheet_parti["A1"].value==0:
                    await client.send_message(channel,'종료할 투표가 존재하지 않습니다.')
                    flag=False

                if flag:
                    vote=[]

                    n=sheet_vote["A1"].value
                    
                    for i in range(2,2+n):
                        vote.append([sheet_vote["B"+str(i)].value,sheet_vote["A"+str(i)].value])

                    vote.sort(reverse=True)

                    des=''
                    tmp1=0
                    tmp2=0
                    for i in range(0,n):
                        if tmp1!=vote[i][0]:
                            tmp2=i+1
                            tmp1=vote[i][0]
                        des=des+"%d"%tmp2+'위 - <'+vote[i][1]+'> --------------- '+"%d"%vote[i][0]+"표"+'\n'+'\n'
                    
                    embed=discord.Embed(title="이전투표결과",description=des, color=0xff0000)
                    embed.set_footer(text='투표 총 참여자 - '+"%d"%sheet_parti["B1"].value+"명")
                    await client.send_message(channel,embed=embed)
                    
            #참여자 목록 초기화
            sheet_parti["B1"]=0
            for i in range(2,32):
                sheet_parti["A"+str(i)].value='-'
                sheet_parti["B"+str(i)].value=0
            file_parti.save("참여자.xlsx")
            
            tmp=message.content[8:]
            candidate=tmp.split('/')

            #투표 목록 초기화
            sheet_vote["A1"]=len(candidate)

            cnt=2
            for i in candidate:
                sheet_vote["A"+str(cnt)].value=i
                sheet_vote["B"+str(cnt)].value=0
                cnt+=1

            file_vote.save("투표목록.xlsx")
            
            #투표 목록 출력
            des=''
            cnt=1
            for i in candidate:
                des=des+"%d"%cnt+'번째 후보 : '+i+'\n'+'\n'
                cnt+=1
            embed=discord.Embed(title="후보",description=des, color=0x00ff00)
            embed.set_image(url="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT-4pFRHcvahEWCQkH-horx1XFydmx3w3ZtNn2wwfOu5KB5IxfY")
            embed.set_footer(text='우덜식 민주주의')
            
            await client.send_message(channel,embed=embed)
            
            

    #투표참여
    elif message.content.startswith('!투표참여'):
        file=openpyxl.load_workbook("참여자.xlsx")
        file_vote=openpyxl.load_workbook("투표목록.xlsx")
        file_private=openpyxl.load_workbook("비공개투표.xlsx")

        sheet=file.active
        sheet_vote=file_vote.active
        sheet_private=file_private.active

        per=sheet["B1"].value+1
        
        if sheet["A1"].value==0:
            await client.send_message(channel,'참여할 투표가 존재하지 않습니다.')
        else:
            cnt_member=1
            flag=True
            for i in  range(2,2+per):
                if str(sheet["A"+str(i)].value)=='-':
                    sheet["A"+str(i)].value=str(id)
                    break
                elif str(sheet["A"+str(i)].value)==str(id):
                    await client.send_message(channel,'이미 투표에 참여되었습니다')
                    flag=False
                    break
                
            if flag:
                sheet["B1"].value+=1
                n_parti=sheet["B1"].value
                n_vote=sheet_vote["A1"].value
                file.save("참여자.xlsx")
                await client.send_message(channel,'<@'+id+'>님이 투표에 참여하셨습니다.')

                if sheet["A1"].value==2:
                    t=random.randint(0,n_vote)
                    for i in range(1,n_vote+1):
                        sheet_private.cell(n_parti,i).value=((i+t)%n_vote)+1
                    file_private.save("비공개투표.xlsx")
                    await client.send_message(channel,'암호화 된 번호를 보냈습니다.')
                    me=await client.get_user_info(id)

                    des='암호화 된 번호입니다.\n\n'
                    for i in range(1,n_vote+1):
                        des=des+'후보 <'+str(sheet_vote["A"+str(i+1)].value)+'> ---------- '+str(sheet_private.cell(n_parti,i).value)+'번\n\n'
                    embed=discord.Embed(title="암호화 된 번호 목록",description=des, color=0x00ffff)
                    await client.send_message(me,embed=embed)
           
                

    #투표종료
    elif message.content.startswith('!투표종료'):
        file=openpyxl.load_workbook("투표목록.xlsx")
        sheet=file.active
        
        file_parti=openpyxl.load_workbook("참여자.xlsx")
        sheet_parti=file_parti.active

        flag=True
        if sheet_parti["A1"].value==0:
            await client.send_message(channel,'종료할 투표가 존재하지 않습니다.')
            flag=False

        if flag:
            vote=[]

            n=sheet["A1"].value
            
            for i in range(2,2+n):
                vote.append([sheet["B"+str(i)].value,sheet["A"+str(i)].value])

            vote.sort(reverse=True)

            des=''
            tmp1=0
            tmp2=0
            for i in range(0,n):
                if tmp1!=vote[i][0]:
                    tmp2=i+1
                    tmp1=vote[i][0]
                des=des+"%d"%tmp2+'위 - <'+vote[i][1]+'> --------------- '+"%d"%vote[i][0]+"표"+'\n'+'\n'

            embed=discord.Embed(title="투표결과",description=des, color=0xff0000)
            embed.set_footer(text='투표 총 참여자 - '+"%d"%sheet_parti["B1"].value+"명")
            await client.send_message(channel,embed=embed)

            #참여자 목록에 투표종료를 저장
            sheet_parti["A1"].value=0
            file_parti.save("참여자.xlsx")
            

    #투표
    elif message.content.startswith('!투표'):
        file_parti=openpyxl.load_workbook("참여자.xlsx")
        file_vote=openpyxl.load_workbook("투표목록.xlsx")
        file_private=openpyxl.load_workbook("비공개투표.xlsx")

        sheet_parti=file_parti.active
        sheet_vote=file_vote.active
        sheet_private=file_private.active

        flag1=True
        if sheet_parti["A1"].value==0:
            await client.send_message(channel,'참여할 투표가 존재하지 않습니다.')
            flag1=False
        elif message.content[3]!=' ':
            await client.send_message(channel,'잘못된 입력 형식입니다.')
            flag1=False

        if flag1:    
            n=sheet_vote["A1"].value
            per=sheet_parti["B1"].value
            idx=int(message.content[4:])#사용자가 투표한 번호

            flag=False
            t=0
            idx_parti=0
            for i in range(2,2+per):
                if str(sheet_parti["A"+str(i)].value)==str(id):
                      flag=True
                      t=sheet_parti["B"+str(i)].value
                      idx_parti=i
                      break


            if flag:
                if idx<=0 or idx>n:
                    await client.send_message(channel, '없는 후보입니다.')
                else:
                    if t!=0:
                        sheet_vote["B"+str(t+1)].value-=1
                        
                    if sheet_parti["A1"].value==2:
                        for i in range(1,n+1):
                            if sheet_private.cell(idx_parti-1,i).value==idx:
                                idx=i
                                break
                    
                    sheet_vote["B"+str(idx+1)].value+=1
                    sheet_parti["B"+str(idx_parti)].value=idx

                    file_parti.save("참여자.xlsx")
                    file_vote.save("투표목록.xlsx")
                    await client.send_message(channel,'투표 완료')
            else:
                await client.send_message(channel,'투표에 참여하지 않았습니다. 투표에 참여 후 다시 시도해주세요.')
            
client.run(token)
